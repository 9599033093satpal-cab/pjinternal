"""
Aether OCR — Excel Exporter
==============================
Exports master_case.json to structured Excel workbook.
Sheets: Summary, Pages, Fields, Audit Trail
"""

import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Run: pip install openpyxl")

# Color palette
COLOR_HEADER     = "1A237E"   # Deep navy
COLOR_HEADER_TXT = "FFFFFF"
COLOR_GREEN      = "C8E6C9"   # Light green — high confidence
COLOR_YELLOW     = "FFF9C4"   # Light yellow — needs review
COLOR_RED        = "FFCDD2"   # Light red — must correct
COLOR_ALT_ROW    = "F5F5F5"   # Alternating row


def export_to_excel(master_case: dict, output_dir: str) -> str:
    """
    Export full master_case.json to Excel.
    Returns path to saved .xlsx file.
    """
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl not installed")
        return ""

    wb = openpyxl.Workbook()

    # ── Single Sheet: OCR Text ──
    ws = wb.active
    ws.title = "OCR Text"
    _write_raw_text_sheet(ws, master_case)

    # Save
    case_id = master_case.get("case_id", "case")[:8]
    filename = f"aether_export_{case_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = Path(output_dir) / filename
    wb.save(str(output_path))
    logger.info(f"Excel exported: {output_path}")
    return str(output_path)


def _write_summary_sheet(ws, case: dict):
    """Sheet 1: High-level case summary."""
    _set_header_row(ws, ["Field", "Value", "Confidence", "Status"], row=1)

    rows = []

    # Document info
    rows += [
        ("Case ID", case.get("case_id", ""), "", ""),
        ("Filename", case.get("filename", ""), "", ""),
        ("Created At", case.get("created_at", ""), "", ""),
        ("Primary Doc Type", case.get("document_bundle", {}).get("primary_type", ""), "", ""),
        ("Total Pages", case.get("document_bundle", {}).get("total_pages", ""), "", ""),
        ("Blank Pages", case.get("document_bundle", {}).get("blank_pages", ""), "", ""),
        ("Overall Confidence", case.get("overall_confidence", ""), "", case.get("confidence_status", "")),
        ("", "", "", ""),
        # Financial
        ("FINANCIAL", "", "", ""),
        ("Loan Number", case.get("financial", {}).get("loan_number", ""), _get_conf(case, "financial.loan_number"), _get_status(case, "financial.loan_number")),
        ("Loan Amount", case.get("financial", {}).get("loan_amount", ""), _get_conf(case, "financial.loan_amount"), _get_status(case, "financial.loan_amount")),
        ("Outstanding Principal", case.get("financial", {}).get("outstanding_principal", ""), _get_conf(case, "financial.outstanding_principal"), ""),
        ("Outstanding Interest", case.get("financial", {}).get("outstanding_interest", ""), "", ""),
        ("Total Demand", case.get("financial", {}).get("total_demand", ""), _get_conf(case, "financial.total_demand"), ""),
        ("Interest Rate", case.get("financial", {}).get("interest_rate", ""), "", ""),
        ("Tenure (months)", case.get("financial", {}).get("tenure_months", ""), "", ""),
        ("EMI Amount", case.get("financial", {}).get("emi_amount", ""), "", ""),
        ("", "", "", ""),
        # Parties
        ("PARTIES", "", "", ""),
        ("Borrower Name", case.get("parties", {}).get("borrower", {}).get("name", ""), _get_conf(case, "parties.borrower.name"), ""),
        ("Borrower Address", case.get("parties", {}).get("borrower", {}).get("address", ""), "", ""),
        ("Co-Borrower", case.get("parties", {}).get("co_borrower", {}).get("name", ""), "", ""),
        ("Guarantor", case.get("parties", {}).get("guarantor", {}).get("name", ""), "", ""),
        ("Bank Name", case.get("parties", {}).get("lender", {}).get("bank_name", ""), "", ""),
        ("Branch", case.get("parties", {}).get("lender", {}).get("branch", ""), "", ""),
        ("Authorized Officer", case.get("parties", {}).get("lender", {}).get("officer", ""), "", ""),
        ("", "", "", ""),
        # Property
        ("PROPERTY", "", "", ""),
        ("Survey No", case.get("property", {}).get("survey_number", ""), "", ""),
        ("Area", case.get("property", {}).get("area", "") + " " + case.get("property", {}).get("area_unit", ""), "", ""),
        ("Address", case.get("property", {}).get("address", ""), "", ""),
        ("District", case.get("property", {}).get("district", ""), "", ""),
        ("State", case.get("property", {}).get("state", ""), "", ""),
        ("", "", "", ""),
        # Legal
        ("LEGAL STATUS", "", "", ""),
        ("Notice Date", case.get("legal_status", {}).get("notice_date", ""), "", ""),
        ("Response Deadline", case.get("legal_status", {}).get("response_deadline", ""), "", ""),
        ("Court Case No.", case.get("legal_status", {}).get("court_case_number", ""), "", ""),
    ]

    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=str(val) if val else "")
            # Section headers bold
            if j == 1 and str(val).isupper() and val:
                cell.font = Font(bold=True, color="1A237E")
            # Alternating rows
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
            # Confidence coloring
            if j == 3 and val:
                try:
                    score = float(val)
                    if score >= 0.85:
                        cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
                    elif score >= 0.60:
                        cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)
                    else:
                        cell.fill = PatternFill("solid", fgColor=COLOR_RED)
                except ValueError:
                    pass

    _auto_width(ws)


def _write_page_sheet(ws, case: dict):
    """Sheet 2: Per-page OCR results."""
    _set_header_row(ws, ["Page", "Doc Type", "OCR Engine", "Quality", "Confidence", "Is Blank", "Text Length"], row=1)

    for i, page in enumerate(case.get("pages", []), start=2):
        row = [
            page.get("page_num", ""),
            page.get("doc_type", ""),
            page.get("ocr_engine", ""),
            page.get("quality_score", ""),
            page.get("confidence", ""),
            "Yes" if page.get("is_blank") else "No",
            len(str(page.get("extracted_fields", {}))),
        ]
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            # Color by confidence
            if j == 5 and isinstance(val, float):
                if val >= 0.85:
                    cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
                elif val >= 0.60:
                    cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)
                else:
                    cell.fill = PatternFill("solid", fgColor=COLOR_RED)

    _auto_width(ws)


def _write_fields_sheet(ws, case: dict):
    """Sheet 3: All extracted fields with source pages."""
    _set_header_row(ws, ["Field Path", "Value", "Confidence Score", "Status", "Color", "Source Page"], row=1)

    conf_map = case.get("confidence_map", {})
    i = 2
    for field_path, conf_data in conf_map.items():
        # Get value from master case
        value = _get_nested_value(case, field_path)
        ws.cell(row=i, column=1, value=field_path)
        ws.cell(row=i, column=2, value=str(value) if value else "")
        ws.cell(row=i, column=3, value=conf_data.get("score", ""))
        ws.cell(row=i, column=4, value=conf_data.get("status", ""))
        color_cell = ws.cell(row=i, column=5, value=conf_data.get("color", ""))
        # Fill color
        color_map = {"green": COLOR_GREEN, "yellow": COLOR_YELLOW, "red": COLOR_RED}
        bg = color_map.get(conf_data.get("color", ""), "FFFFFF")
        color_cell.fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=i, column=6, value="")
        i += 1

    _auto_width(ws)


def _write_audit_sheet(ws, case: dict):
    """Sheet 4: Human correction audit trail."""
    _set_header_row(ws, ["Field", "Machine Value", "Human Value", "Status", "Edit #", "Accuracy %"], row=1)

    audit = case.get("audit", {})
    trail = audit.get("trail", [])

    for i, entry in enumerate(trail, start=2):
        ws.cell(row=i, column=1, value=entry.get("field", ""))
        ws.cell(row=i, column=2, value=str(entry.get("machine", "")))
        ws.cell(row=i, column=3, value=str(entry.get("human", "")))
        status_cell = ws.cell(row=i, column=4, value=entry.get("status", ""))
        ws.cell(row=i, column=5, value=audit.get("edit_count", ""))
        ws.cell(row=i, column=6, value=entry.get("accuracy", ""))

        if entry.get("status") == "matched":
            status_cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
        else:
            status_cell.fill = PatternFill("solid", fgColor=COLOR_RED)

    _auto_width(ws)


def _set_header_row(ws, headers, row=1):
    for j, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=j, value=header)
        cell.font = Font(bold=True, color=COLOR_HEADER_TXT, size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 20


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def _get_conf(case, path):
    return case.get("confidence_map", {}).get(path, {}).get("score", "")


def _get_status(case, path):
    return case.get("confidence_map", {}).get(path, {}).get("status", "")


def _get_nested_value(obj, path):
    keys = path.split(".")
    for k in keys:
        if isinstance(obj, dict):
            obj = obj.get(k, "")
        else:
            return ""
    return obj


def _write_raw_text_sheet(ws, case: dict):
    """Sheet 5: Raw OCR text page-by-page."""
    _set_header_row(ws, ["Page", "Raw OCR Text"], row=1)
    
    wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    for i, page in enumerate(case.get("pages", []), start=2):
        ws.cell(row=i, column=1, value=f"Page {page.get('page_num', '')}")
        cell_text = ws.cell(row=i, column=2, value=page.get("text", ""))
        cell_text.alignment = wrap_alignment
        
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 100
